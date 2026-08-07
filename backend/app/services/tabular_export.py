import csv
import io

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

DEFAULT_WIDTH = 24
WIDE_WIDTH = 60


class TabularExporter:
    def __init__(
        self, columns: list[str], sheet_title: str, wide_columns: tuple[str, ...] = ()
    ) -> None:
        self.columns = columns
        self.sheet_title = sheet_title
        self.wide_columns = set(wide_columns)

    def to_csv(self, rows: list[dict[str, str]]) -> str:
        buffer = io.StringIO()
        buffer.write("﻿")
        writer = csv.DictWriter(buffer, fieldnames=self.columns)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)
        return buffer.getvalue()

    def to_xlsx(self, rows: list[dict[str, str]]) -> bytes:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = self.sheet_title

        sheet.append(self.columns)
        for cell in sheet[1]:
            cell.font = Font(bold=True)

        for row in rows:
            sheet.append([row.get(column, "") for column in self.columns])

        for index, column in enumerate(self.columns, start=1):
            width = WIDE_WIDTH if column in self.wide_columns else DEFAULT_WIDTH
            sheet.column_dimensions[get_column_letter(index)].width = width
        sheet.freeze_panes = "A2"

        buffer = io.BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()
